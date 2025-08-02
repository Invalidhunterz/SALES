from flask import Flask, render_template, request
from openpyxl import load_workbook

app = Flask(__name__)
EXCEL_PATH = 'your_file.xlsx'  # Change to your Excel filename

@app.route('/', methods=['GET'])
def input_form():
    return render_template('input.html')

@app.route('/save', methods=['POST'])
def save_to_excel():
    firm = request.form.get('firm')
    data = request.form.get('data')
    period = request.form.get('period')
    gstno = request.form.get('gstno')

    if not all([firm, data, period, gstno]):
        return render_template('input.html', message="All fields are required.")

    wb = load_workbook(EXCEL_PATH)
    ws = wb.active

    ws['C2'] = firm
    ws['C3'] = data
    ws['C4'] = period
    ws['C5'] = gstno

    wb.save(EXCEL_PATH)
    return render_template('input.html', message="Saved all values to Excel!")

if __name__ == '__main__':
    app.run(debug=True)
